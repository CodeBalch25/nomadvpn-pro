from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = Workbook()

# Colors
header_fill = PatternFill('solid', fgColor='1E3A5F')
header_font = Font(bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill('solid', fgColor='2E5077')
highlight_fill = PatternFill('solid', fgColor='E8F4FD')
success_fill = PatternFill('solid', fgColor='D4EDDA')
warning_fill = PatternFill('solid', fgColor='FFF3CD')
input_font = Font(color='0000FF')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(cell):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def style_subheader(cell):
    cell.fill = subheader_fill
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def style_cell(cell, is_input=False):
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if is_input:
        cell.font = input_font

# ============ SHEET 1: DASHBOARD ============
ws = wb.active
ws.title = "Dashboard"

ws['A1'] = "NomadVPN Pro - Business Tracker"
ws['A1'].font = Font(bold=True, size=18, color='1E3A5F')
ws.merge_cells('A1:F1')

ws['A3'] = "KEY METRICS"
style_header(ws['A3'])
ws.merge_cells('A3:B3')

metrics = [
    ("Total Revenue", "=Customers!H2", "$"),
    ("Total Customers", "=COUNTA(Customers!A:A)-1", ""),
    ("Email Subscribers", "", ""),
    ("Website Visitors", "", ""),
    ("Consultations Booked", "", ""),
    ("Conversion Rate", "=IF(E8>0,E6/E8,0)", "%"),
]

for i, (label, formula, fmt) in enumerate(metrics, start=4):
    ws[f'A{i}'] = label
    ws[f'B{i}'] = formula if formula else 0
    style_cell(ws[f'A{i}'])
    style_cell(ws[f'B{i}'], is_input=not formula)
    if fmt == "$":
        ws[f'B{i}'].number_format = '$#,##0'
    elif fmt == "%":
        ws[f'B{i}'].number_format = '0.0%'

ws['D3'] = "MONTHLY TARGETS"
style_header(ws['D3'])
ws.merge_cells('D3:F3')

targets = [
    ("Month 1", "Revenue", 2000),
    ("", "Customers", 3),
    ("", "Email Subs", 50),
    ("Month 3", "Revenue", 7500),
    ("", "Customers", 12),
    ("", "Email Subs", 200),
]

row = 4
for month, metric, target in targets:
    ws[f'D{row}'] = month
    ws[f'E{row}'] = metric
    ws[f'F{row}'] = target
    style_cell(ws[f'D{row}'])
    style_cell(ws[f'E{row}'])
    style_cell(ws[f'F{row}'])
    if "Revenue" in metric:
        ws[f'F{row}'].number_format = '$#,##0'
    row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12

# ============ SHEET 2: 90-DAY MARKETING PLAN ============
ws2 = wb.create_sheet("Marketing Plan")

ws2['A1'] = "90-Day Marketing Launch Plan"
ws2['A1'].font = Font(bold=True, size=16, color='1E3A5F')
ws2.merge_cells('A1:G1')

headers = ["Week", "Phase", "Task", "Channel", "Priority", "Status", "Due Date", "Notes"]
for col, header in enumerate(headers, start=1):
    cell = ws2.cell(row=3, column=col, value=header)
    style_header(cell)

tasks = [
    # Week 1-2: Foundation
    (1, "Foundation", "Apply GL.iNet Affiliate Program", "Partnership", "HIGH", "Not Started", 1),
    (1, "Foundation", "Apply GL.iNet Reseller Program", "Partnership", "HIGH", "Not Started", 1),
    (1, "Foundation", "Create Reddit account (u/NomadVPNPro)", "Reddit", "HIGH", "Not Started", 1),
    (1, "Foundation", "Start commenting on r/digitalnomad (10+ helpful comments)", "Reddit", "HIGH", "Not Started", 7),
    (1, "Foundation", "Start commenting on r/GlInet (5+ helpful comments)", "Reddit", "HIGH", "Not Started", 7),
    (2, "Foundation", "Create lead magnet PDF", "Email", "MEDIUM", "Not Started", 10),
    (2, "Foundation", "Set up email capture (ConvertKit)", "Email", "MEDIUM", "Not Started", 10),
    (2, "Foundation", "Join GL.iNet forum, introduce yourself", "Community", "MEDIUM", "Not Started", 10),
    
    # Week 3-4: Content Launch
    (3, "Content Launch", "Draft first Reddit educational post", "Reddit", "HIGH", "Not Started", 15),
    (3, "Content Launch", "Post to r/GlInet (test audience)", "Reddit", "HIGH", "Not Started", 17),
    (3, "Content Launch", "Script YouTube video: Flint 2 Setup Guide", "YouTube", "HIGH", "Not Started", 18),
    (4, "Content Launch", "Record & publish first YouTube video", "YouTube", "HIGH", "Not Started", 25),
    (4, "Content Launch", "Post educational content to r/digitalnomad", "Reddit", "HIGH", "Not Started", 24),
    (4, "Content Launch", "First blog post live (SEO)", "SEO", "MEDIUM", "Not Started", 28),
    
    # Week 5-8: Amplification
    (5, "Amplification", "Reddit AMA post", "Reddit", "HIGH", "Not Started", 35),
    (6, "Amplification", "Second YouTube video", "YouTube", "MEDIUM", "Not Started", 42),
    (6, "Amplification", "Second blog post", "SEO", "MEDIUM", "Not Started", 42),
    (7, "Amplification", "Third YouTube video", "YouTube", "MEDIUM", "Not Started", 49),
    (7, "Amplification", "Third blog post", "SEO", "MEDIUM", "Not Started", 49),
    (8, "Amplification", "Test Reddit ads ($100-200)", "Paid", "LOW", "Not Started", 56),
    
    # Week 9-12: Optimization
    (9, "Optimization", "Analyze traffic & conversion data", "Analytics", "HIGH", "Not Started", 63),
    (10, "Optimization", "Launch referral program", "Growth", "MEDIUM", "Not Started", 70),
    (11, "Optimization", "Collect video testimonials", "Social Proof", "HIGH", "Not Started", 77),
    (12, "Optimization", "Plan Q2 strategy", "Strategy", "MEDIUM", "Not Started", 84),
]

start_date = datetime.now()
for i, (week, phase, task, channel, priority, status, days) in enumerate(tasks, start=4):
    ws2.cell(row=i, column=1, value=week)
    ws2.cell(row=i, column=2, value=phase)
    ws2.cell(row=i, column=3, value=task)
    ws2.cell(row=i, column=4, value=channel)
    ws2.cell(row=i, column=5, value=priority)
    ws2.cell(row=i, column=6, value=status)
    ws2.cell(row=i, column=7, value=start_date + timedelta(days=days))
    ws2.cell(row=i, column=8, value="")
    
    for col in range(1, 9):
        cell = ws2.cell(row=i, column=col)
        style_cell(cell)
        if col == 6:
            cell.font = input_font
    
    ws2.cell(row=i, column=7).number_format = 'MM/DD/YYYY'
    
    if priority == "HIGH":
        ws2.cell(row=i, column=5).fill = warning_fill

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 30

# ============ SHEET 3: CUSTOMERS ============
ws3 = wb.create_sheet("Customers")

headers = ["Name", "Email", "Phone", "Service Tier", "Price", "Status", "Date", "Revenue", "Notes"]
for col, header in enumerate(headers, start=1):
    cell = ws3.cell(row=1, column=col, value=header)
    style_header(cell)

ws3['H2'] = '=SUMIF(F:F,"Completed",E:E)'
style_cell(ws3['H2'])
ws3['H2'].number_format = '$#,##0'

for col in range(1, 10):
    ws3.column_dimensions[get_column_letter(col)].width = 15

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['I'].width = 30

# ============ SHEET 4: REDDIT TRACKER ============
ws4 = wb.create_sheet("Reddit Tracker")

ws4['A1'] = "Reddit Activity Tracker"
ws4['A1'].font = Font(bold=True, size=14, color='1E3A5F')

headers = ["Date", "Subreddit", "Type", "Title/Topic", "Upvotes", "Comments", "Link", "Leads Generated"]
for col, header in enumerate(headers, start=1):
    cell = ws4.cell(row=3, column=col, value=header)
    style_header(cell)

subreddits = ["r/digitalnomad", "r/GlInet", "r/VPN", "r/remotework", "r/homelab", "r/overemployed"]
ws4['J3'] = "Target Subreddits"
style_header(ws4['J3'])
for i, sub in enumerate(subreddits, start=4):
    ws4[f'J{i}'] = sub
    style_cell(ws4[f'J{i}'])

for col in range(1, 9):
    ws4.column_dimensions[get_column_letter(col)].width = 15
ws4.column_dimensions['D'].width = 40
ws4.column_dimensions['G'].width = 30

# ============ SHEET 5: YOUTUBE TRACKER ============
ws5 = wb.create_sheet("YouTube Tracker")

ws5['A1'] = "YouTube Content Tracker"
ws5['A1'].font = Font(bold=True, size=14, color='1E3A5F')

headers = ["Video Title", "Status", "Publish Date", "Views", "Likes", "Subscribers Gained", "Clicks to Site", "Revenue (Affiliate)"]
for col, header in enumerate(headers, start=1):
    cell = ws5.cell(row=3, column=col, value=header)
    style_header(cell)

video_ideas = [
    ("GL.iNet Flint 2 Complete Setup - WireGuard VPN Server", "Planned"),
    ("Best Travel Router 2025 - GL.iNet Beryl AX Review", "Planned"),
    ("Why Commercial VPNs Fail for Remote Work", "Planned"),
    ("Work From Anywhere Setup Guide - Digital Nomad VPN", "Planned"),
    ("Residential IP vs Data Center IP Explained", "Planned"),
]

for i, (title, status) in enumerate(video_ideas, start=4):
    ws5.cell(row=i, column=1, value=title)
    ws5.cell(row=i, column=2, value=status)
    for col in range(1, 9):
        style_cell(ws5.cell(row=i, column=col))

ws5.column_dimensions['A'].width = 50
for col in range(2, 9):
    ws5.column_dimensions[get_column_letter(col)].width = 15

# ============ SHEET 6: FINANCIAL TRACKER ============
ws6 = wb.create_sheet("Financials")

ws6['A1'] = "Financial Tracker"
ws6['A1'].font = Font(bold=True, size=14, color='1E3A5F')

# Revenue Section
ws6['A3'] = "REVENUE"
style_header(ws6['A3'])
ws6.merge_cells('A3:C3')

revenue_headers = ["Service Tier", "Price", "Units Sold", "Total"]
for col, header in enumerate(revenue_headers, start=1):
    cell = ws6.cell(row=4, column=col, value=header)
    style_subheader(cell)

services = [
    ("Essential Setup", 699, 0),
    ("Premium + Support", 1299, 0),
    ("Remote VPN Access (Monthly)", 45, 0),
]

for i, (service, price, units) in enumerate(services, start=5):
    ws6.cell(row=i, column=1, value=service)
    ws6.cell(row=i, column=2, value=price)
    ws6.cell(row=i, column=3, value=units)
    ws6.cell(row=i, column=4, value=f'=B{i}*C{i}')
    for col in range(1, 5):
        style_cell(ws6.cell(row=i, column=col))
        if col in [2, 4]:
            ws6.cell(row=i, column=col).number_format = '$#,##0'
    ws6.cell(row=i, column=3).font = input_font

ws6['A8'] = "Total Revenue"
ws6['D8'] = '=SUM(D5:D7)'
ws6['A8'].font = Font(bold=True)
ws6['D8'].font = Font(bold=True)
ws6['D8'].number_format = '$#,##0'
style_cell(ws6['A8'])
style_cell(ws6['D8'])

# Expenses Section
ws6['A10'] = "EXPENSES"
style_header(ws6['A10'])
ws6.merge_cells('A10:C10')

expense_headers = ["Category", "Budgeted", "Actual", "Variance"]
for col, header in enumerate(expense_headers, start=1):
    cell = ws6.cell(row=11, column=col, value=header)
    style_subheader(cell)

expenses = [
    ("Hardware Inventory", 800, 0),
    ("Email Marketing", 50, 0),
    ("Video Equipment", 100, 0),
    ("Reddit Ads", 300, 0),
    ("Domain/Hosting", 120, 0),
    ("Miscellaneous", 100, 0),
]

for i, (category, budgeted, actual) in enumerate(expenses, start=12):
    ws6.cell(row=i, column=1, value=category)
    ws6.cell(row=i, column=2, value=budgeted)
    ws6.cell(row=i, column=3, value=actual)
    ws6.cell(row=i, column=4, value=f'=B{i}-C{i}')
    for col in range(1, 5):
        style_cell(ws6.cell(row=i, column=col))
        if col in [2, 3, 4]:
            ws6.cell(row=i, column=col).number_format = '$#,##0'
    ws6.cell(row=i, column=3).font = input_font

ws6['A18'] = "Total Expenses"
ws6['B18'] = '=SUM(B12:B17)'
ws6['C18'] = '=SUM(C12:C17)'
ws6['D18'] = '=B18-C18'
for col in ['A', 'B', 'C', 'D']:
    ws6[f'{col}18'].font = Font(bold=True)
    style_cell(ws6[f'{col}18'])
    if col in ['B', 'C', 'D']:
        ws6[f'{col}18'].number_format = '$#,##0'

# Profit
ws6['A20'] = "NET PROFIT"
ws6['A20'].font = Font(bold=True, size=12)
ws6['D20'] = '=D8-C18'
ws6['D20'].font = Font(bold=True, size=12)
ws6['D20'].number_format = '$#,##0'
style_cell(ws6['A20'])
style_cell(ws6['D20'])
ws6['D20'].fill = success_fill

for col in range(1, 5):
    ws6.column_dimensions[get_column_letter(col)].width = 20

# ============ SHEET 7: PARTNERSHIPS ============
ws7 = wb.create_sheet("Partnerships")

ws7['A1'] = "GL.iNet Partnership Tracker"
ws7['A1'].font = Font(bold=True, size=14, color='1E3A5F')

headers = ["Partnership Type", "Status", "Applied Date", "Approval Date", "Benefits", "Action Items"]
for col, header in enumerate(headers, start=1):
    cell = ws7.cell(row=3, column=col, value=header)
    style_header(cell)

partnerships = [
    ("Affiliate Program", "Not Applied", "", "", "10% commission on sales", "Apply at gl-inet.com/affiliate-program/"),
    ("Reseller Program", "Not Applied", "", "", "15-25% wholesale discount, GoodCloud access, Where to Buy listing", "Apply at gl-inet.com/form/become-a-reseller/"),
    ("Forum Presence", "Not Started", "", "", "Community credibility, lead generation", "Join forum.gl-inet.com, introduce yourself"),
    ("Distributor Program", "Future", "", "", "Regional exclusivity, deeper discounts", "Pursue after $100K+ revenue"),
]

for i, (ptype, status, applied, approved, benefits, actions) in enumerate(partnerships, start=4):
    ws7.cell(row=i, column=1, value=ptype)
    ws7.cell(row=i, column=2, value=status)
    ws7.cell(row=i, column=3, value=applied)
    ws7.cell(row=i, column=4, value=approved)
    ws7.cell(row=i, column=5, value=benefits)
    ws7.cell(row=i, column=6, value=actions)
    for col in range(1, 7):
        style_cell(ws7.cell(row=i, column=col))
    ws7.cell(row=i, column=2).font = input_font

ws7.column_dimensions['A'].width = 20
ws7.column_dimensions['B'].width = 15
ws7.column_dimensions['C'].width = 15
ws7.column_dimensions['D'].width = 15
ws7.column_dimensions['E'].width = 50
ws7.column_dimensions['F'].width = 50

# Save
wb.save('C:/Users/timud/nomadvpn-pro/NomadVPN_Business_Tracker.xlsx')
print("Excel tracker created successfully!")
